import os
import webbrowser

from datetime import datetime

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QCheckBox,
    QPushButton,
    QScrollArea,
    QFrame,
    QSizePolicy
)

from openpyxl import load_workbook

from widget_state import (
    load_completed_items,
    mark_completed,
    unmark_completed
)


EXCEL_FILE = "data/College_Placement_2026.xlsx"

REFRESH_INTERVAL = 30 * 1000  # 30 seconds


class MailItem(QFrame):

    def __init__(
        self,
        message_id,
        summary,
        link,
        on_complete
    ):
        super().__init__()

        self.message_id = message_id
        self.link = link
        self.on_complete = on_complete

        self.setFrameShape(QFrame.StyledPanel)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(5)

        # Checkbox
        self.checkbox = QCheckBox(summary)

        self.checkbox.setFont(
            QFont("Segoe UI", 10)
        )

        self.checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
            }

            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)

        self.checkbox.stateChanged.connect(
            self.handle_checkbox
        )

        layout.addWidget(self.checkbox)

        # Open mail button
        if link:

            self.mail_button = QPushButton(
                "Open Mail"
            )

            self.mail_button.setCursor(
                Qt.PointingHandCursor
            )

            self.mail_button.clicked.connect(
                self.open_mail
            )

            layout.addWidget(
                self.mail_button
            )

    def handle_checkbox(self, state):

        if state == Qt.Checked:

            mark_completed(
                self.message_id
            )

            self.on_complete()

        else:

            unmark_completed(
                self.message_id
            )

    def open_mail(self):

        if self.link:
            webbrowser.open(self.link)


class MailWidget(QWidget):

    def __init__(self):

        super().__init__()

        self.setWindowTitle(
            "College Mail"
        )

        self.resize(900, 650)

        self.setup_ui()

        # Initial load
        self.refresh()

        # Automatic refresh
        self.timer = QTimer(self)

        self.timer.timeout.connect(
            self.refresh
        )

        self.timer.start(
            REFRESH_INTERVAL
        )

    # --------------------------------------------------
    # UI
    # --------------------------------------------------

    def setup_ui(self):

        main_layout = QVBoxLayout(self)

        main_layout.setContentsMargins(
            15, 15, 15, 15
        )

        # Header
        header = QHBoxLayout()

        title = QLabel(
            "📬 College Mail"
        )

        title.setFont(
            QFont("Segoe UI", 18, QFont.Bold)
        )

        header.addWidget(title)

        header.addStretch()

        self.updated_label = QLabel(
            "Updating..."
        )

        header.addWidget(
            self.updated_label
        )

        refresh_button = QPushButton(
            "⟳"
        )

        refresh_button.setFixedWidth(40)

        refresh_button.clicked.connect(
            self.refresh
        )

        header.addWidget(
            refresh_button
        )

        main_layout.addLayout(header)

        # Two columns
        columns = QHBoxLayout()

        # Placement
        placement_widget = self.create_section(
            "PLACEMENT"
        )

        # Academic
        academic_widget = self.create_section(
            "ACADEMIC"
        )

        self.placement_layout = (
            placement_widget.layout()
        )

        self.academic_layout = (
            academic_widget.layout()
        )

        columns.addWidget(
            placement_widget
        )

        columns.addWidget(
            academic_widget
        )

        main_layout.addLayout(
            columns
        )

    # --------------------------------------------------

    def create_section(self, title):

        frame = QFrame()

        frame.setFrameShape(
            QFrame.StyledPanel
        )

        layout = QVBoxLayout(frame)

        title_label = QLabel(title)

        title_label.setFont(
            QFont(
                "Segoe UI",
                13,
                QFont.Bold
            )
        )

        layout.addWidget(
            title_label
        )

        layout.addStretch()

        return frame

    # --------------------------------------------------
    # Excel
    # --------------------------------------------------

    def read_excel(self):

        if not os.path.exists(EXCEL_FILE):
            return []

        try:

            wb = load_workbook(
                EXCEL_FILE,
                read_only=True,
                data_only=True
            )

            emails = []

            for sheet_name in wb.sheetnames:

                ws = wb[sheet_name]

                rows = list(
                    ws.iter_rows(
                        values_only=True
                    )
                )

                if not rows:
                    continue

                headers = rows[0]

                header_index = {
                    header: index
                    for index, header
                    in enumerate(headers)
                    if header
                }

                for row in rows[1:]:

                    if not row:
                        continue

                    def get(column):

                        index = header_index.get(
                            column
                        )

                        if index is None:
                            return ""

                        if index >= len(row):
                            return ""

                        return row[index] or ""

                    message_id = get(
                        "Email ID"
                    )

                    summary = get(
                        "Summary"
                    )

                    link = get(
                        "Gmail Link"
                    )

                    subtype = get(
                        "Placement Type"
                    )

                    if not message_id:
                        continue

                    if not summary:
                        continue

                    emails.append({
                        "id": str(message_id),
                        "summary": str(summary),
                        "link": str(link),
                        "subtype": str(subtype)
                    })

            wb.close()

            return emails

        except PermissionError:

            print(
                "Excel is currently open."
            )

            return []

        except Exception as e:

            print(
                f"Widget Excel error: {e}"
            )

            return []

    # --------------------------------------------------
    # Refresh
    # --------------------------------------------------

    def refresh(self):

        completed = load_completed_items()

        emails = self.read_excel()

        # Remove completed items
        active_emails = [
            email
            for email in emails
            if email["id"] not in completed
        ]

        self.clear_layout(
            self.placement_layout
        )

        self.clear_layout(
            self.academic_layout
        )

        # Placement
        placement = [
            email
            for email in active_emails
            if self.is_placement(email)
        ]

        # Academic
        academic = [
            email
            for email in active_emails
            if not self.is_placement(email)
        ]

        # Sort newest-ish by Excel order
        placement.reverse()
        academic.reverse()

        self.populate_placement(
            placement
        )

        self.populate_academic(
            academic
        )

        self.updated_label.setText(
            "Updated "
            + datetime.now().strftime("%H:%M:%S")
        )

    # --------------------------------------------------

    def is_placement(self, email):

        subtype = email["subtype"].upper()

        return subtype in {
            "NEW_COMPANY",
            "SHORTLIST_TEST",
            "SELECTION",
            "TEST+SHORTLIST",
            "TEST + SHORTLIST"
        }

    # --------------------------------------------------

    def populate_placement(self, emails):

        groups = [
            ("NEW COMPANY", "NEW_COMPANY"),
            ("TEST + SHORTLIST", "SHORTLIST_TEST"),
            ("SELECTION", "SELECTION")
        ]

        for title, subtype in groups:

            matching = [
                email
                for email in emails
                if email["subtype"].upper()
                == subtype
            ]

            if not matching:
                continue

            heading = QLabel(title)

            heading.setFont(
                QFont(
                    "Segoe UI",
                    10,
                    QFont.Bold
                )
            )

            self.placement_layout.insertWidget(
                self.placement_layout.count() - 1,
                heading
            )

            for email in matching:

                item = MailItem(
                    email["id"],
                    email["summary"],
                    email["link"],
                    self.refresh
                )

                self.placement_layout.insertWidget(
                    self.placement_layout.count() - 1,
                    item
                )

    # --------------------------------------------------

    def populate_academic(self, emails):

        if not emails:

            label = QLabel(
                "No pending academic updates."
            )

            label.setWordWrap(True)

            self.academic_layout.insertWidget(
                1,
                label
            )

            return

        for email in emails:

            item = MailItem(
                email["id"],
                email["summary"],
                email["link"],
                self.refresh
            )

            self.academic_layout.insertWidget(
                self.academic_layout.count() - 1,
                item
            )

    # --------------------------------------------------

    def clear_layout(self, layout):

        while layout.count() > 1:

            item = layout.takeAt(1)

            widget = item.widget()

            if widget:
                widget.deleteLater()


if __name__ == "__main__":

    app = QApplication([])

    widget = MailWidget()

    widget.show()

    app.exec()