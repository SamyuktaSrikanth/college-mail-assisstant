import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


EXCEL_FILE = "data/College_Placement_{year}.xlsx"


COLUMNS = [
    "Date",
    "Placement Type",
    "Company Name",
    "Batch",
    "Job Title(s)",
    "Job Location(s)",
    "CTC",
    "Intern Stipend",
    "Eligible Branches",
    "Registration Deadline",
    "Shortlisted Count",
    "Selected Count",
    "Test Date",
    "Test Time",
    "Reporting Time",
    "Venue",
    "Campus",
    "Action Required",
    "Summary",
    "Email ID",
    "Gmail Link"
]


def get_excel_path():
    current_year = datetime.now().year
    return EXCEL_FILE.format(year=current_year)


def get_current_month():
    return datetime.now().strftime("%B")


def init_excel():
    path = get_excel_path()

    # Make sure data folder exists
    os.makedirs(os.path.dirname(path), exist_ok=True)

    if not os.path.exists(path):
        wb = Workbook()

        # Rename default sheet to current month
        ws = wb.active
        ws.title = get_current_month()

        # Add headers
        ws.append(COLUMNS)

        wb.save(path)

    else:
        wb = load_workbook(path)

        current_month = get_current_month()

        # Create current month's sheet if it doesn't exist
        if current_month not in wb.sheetnames:
            ws = wb.create_sheet(title=current_month)
            ws.append(COLUMNS)
            wb.save(path)

        wb.close()


def message_id_exists(message_id):
    """
    Checks whether a Gmail message ID has already
    been processed and stored in the Excel workbook.
    """

    path = get_excel_path()

    if not os.path.exists(path):
        return False

    wb = load_workbook(path, read_only=True)

    try:
        email_id_col = COLUMNS.index("Email ID") + 1

        for sheetname in wb.sheetnames:

            ws = wb[sheetname]

            for row in range(2, ws.max_row + 1):

                cell_value = ws.cell(
                    row=row,
                    column=email_id_col
                ).value

                if cell_value == message_id:
                    return True

        return False

    finally:
        wb.close()


def make_excel_value(value):
    """
    Converts Python values into values that Excel/openpyxl
    can safely store.

    Handles:
    - None
    - lists
    - dictionaries
    - numbers
    - strings
    - booleans
    """

    if value is None:
        return ""

    # Example:
    # ["Mechanical intern", "MBD"]
    #
    # becomes:
    # "Mechanical intern; MBD"
    if isinstance(value, list):
        return "; ".join(
            make_excel_value(item)
            for item in value
        )

    # Example:
    # {"city": "Chennai", "country": "India"}
    #
    # becomes:
    # "city: Chennai; country: India"
    if isinstance(value, dict):
        return "; ".join(
            f"{key}: {make_excel_value(val)}"
            for key, val in value.items()
        )

    # Keep normal Excel-compatible values as they are
    if isinstance(value, (str, int, float, bool)):
        return value

    # Anything unexpected → convert to text
    return str(value)


def append_to_excel(extracted_data, email_metadata):
    """
    Appends extracted email data and metadata
    to the current month's Excel sheet.
    """

    init_excel()

    path = get_excel_path()

    wb = load_workbook(path)

    try:
        current_month = get_current_month()
        ws = wb[current_month]

        row_data = []

        for col in COLUMNS:

            # -------------------------
            # Email metadata
            # -------------------------

            if col == "Date":
                value = email_metadata.get("date", "")

            elif col == "Placement Type":
                value = email_metadata.get("subtype", "")

            elif col == "Email ID":
                value = email_metadata.get("id", "")

            elif col == "Gmail Link":
                value = email_metadata.get("link", "")

            elif col == "Summary":
                value = email_metadata.get("summary", "")

            # -------------------------
            # Extracted placement data
            # -------------------------

            else:
                value = extracted_data.get(col, "")

            # -------------------------
            # Make value Excel-safe
            # -------------------------

            value = make_excel_value(value)

            row_data.append(value)

        # Add complete row
        ws.append(row_data)

        wb.save(path)

    finally:
        wb.close()