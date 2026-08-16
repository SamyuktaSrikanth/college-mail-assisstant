from pathlib import Path
from openpyxl import load_workbook
from email.message import EmailMessage
import base64

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build


# =========================================================
# CONFIG
# =========================================================

PROJECT_DIR = Path(__file__).resolve().parent

EXCEL_FILE = (
    PROJECT_DIR
    / "data"
    / "College_Placement_2026.xlsx"
)

MY_EMAIL = "samyukta.srikanth2023@vitstudent.ac.in"

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.send",
]


# =========================================================
# GMAIL AUTHENTICATION
# =========================================================

def get_gmail_service():

    creds = None

    token_file = PROJECT_DIR / "credentials" / "token.json"
    credentials_file = PROJECT_DIR / "credentials" / "credentials.json"

    # Load existing token
    if token_file.exists():

        try:

            creds = Credentials.from_authorized_user_file(
                str(token_file),
                SCOPES
            )

        except Exception:

            creds = None

    # Refresh or authorize
    if not creds or not creds.valid:

        if creds and creds.expired and creds.refresh_token:

            print("Refreshing Gmail access token...")

            creds.refresh(Request())

        else:

            print("Starting Gmail authorization...")

            flow = InstalledAppFlow.from_client_secrets_file(
                str(credentials_file),
                SCOPES
            )

            creds = flow.run_local_server(
                port=0
            )

        # Save token
        token_file.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with open(
            token_file,
            "w"
        ) as token:

            token.write(
                creds.to_json()
            )

    return build(
        "gmail",
        "v1",
        credentials=creds
    )


# =========================================================
# EXCEL READER
# =========================================================

def read_excel():

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    records = []

    # Read all monthly sheets
    for sheet in workbook.worksheets:

        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        headers = [
            str(value).strip()
            if value is not None
            else ""
            for value in rows[0]
        ]

        for row in rows[1:]:

            # Ignore completely empty rows
            if not any(
                value is not None
                for value in row
            ):
                continue

            record = dict(
                zip(headers, row)
            )

            records.append(record)

    return records


# =========================================================
# CLASSIFICATION
# =========================================================

def classify(record):

    placement_type = str(
        record.get("Placement Type", "")
    ).strip().upper()

    # NEW COMPANY
    if placement_type == "NEW_COMPANY":

        return "NEW COMPANY"

    # TEST + SHORTLIST
    elif placement_type in {
        "SHORTLIST_TEST",
        "SHORTLISTED_TEST",
        "UPDATE"
    }:

        return "TEST + SHORTLIST"

    # SELECTION
    elif placement_type == "SELECTION":

        return "SELECTION"

    # OTHER
    elif placement_type == "OTHER":

        return "OTHER"

    return None


# =========================================================
# GROUP RECORDS
# =========================================================

def group_records(records):

    grouped = {
        "NEW COMPANY": [],
        "TEST + SHORTLIST": [],
        "SELECTION": [],
        "OTHER": []
    }

    for record in records:

        category = classify(record)

        if category is None:
            continue

        grouped[category].append(record)

    return grouped


# =========================================================
# HTML HELPERS
# =========================================================

def html_escape(text):

    """
    Prevent special characters in Excel data
    from breaking the HTML email.
    """

    if text is None:
        return ""

    text = str(text)

    return (
        text
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


# =========================================================
# CREATE HTML EMAIL
# =========================================================

def create_html(grouped):

    html = """
    <!DOCTYPE html>

    <html>

    <body style="
        margin: 0;
        padding: 0;
        background-color: #f5f5f5;
        font-family: Arial, Helvetica, sans-serif;
        color: #333333;
    ">

    <div style="
        max-width: 800px;
        margin: 30px auto;
        background-color: #ffffff;
        padding: 30px;
        border-radius: 12px;
    ">

        <h1 style="
            margin-top: 0;
            margin-bottom: 8px;
            font-size: 24px;
            color: #333333;
        ">
            College Mail Summary
        </h1>

        <p style="
            color: #888888;
            font-size: 13px;
            margin-top: 0;
        ">
            Automatically generated from your college emails
        </p>
    """

    # =====================================================
    # PLACEMENT
    # =====================================================

    placement_categories = [
        "NEW COMPANY",
        "TEST + SHORTLIST",
        "SELECTION"
    ]

    has_placement = any(
        grouped[category]
        for category in placement_categories
    )

    if has_placement:

        html += """
        <h2 style="
            margin-top: 30px;
            padding-bottom: 8px;
            border-bottom: 2px solid #315f3a;
            color: #315f3a;
            font-size: 19px;
        ">
            PLACEMENT
        </h2>
        """

        for category in placement_categories:

            items = grouped[category]

            if not items:
                continue

            html += f"""
            <h3 style="
                margin-top: 22px;
                margin-bottom: 10px;
                color: #4d6f52;
                font-size: 15px;
            ">
                {category}
            </h3>
            """

            for item in items:

                summary = html_escape(
                    item.get("Summary", "")
                ).strip()

                link = html_escape(
                    item.get("Gmail Link", "")
                ).strip()

                # Skip completely empty records
                if not summary and not link:
                    continue

                html += f"""
                <div style="
                    border: 1px solid #dddddd;
                    border-radius: 8px;
                    padding: 14px 16px;
                    margin-bottom: 10px;
                    background-color: #fafafa;
                ">

                    <div style="
                        font-size: 14px;
                        line-height: 1.55;
                    ">
                        {summary}
                    </div>

                    <div style="
                        margin-top: 9px;
                    ">
                        <a
                            href="{link}"
                            style="
                                color: #315f3a;
                                font-size: 12px;
                                font-weight: bold;
                                text-decoration: none;
                            "
                        >
                            OPEN MAIL →
                        </a>
                    </div>

                </div>
                """


    # =====================================================
    # OTHER
    # =====================================================

    if grouped["OTHER"]:

        html += """
        <h2 style="
            margin-top: 30px;
            padding-bottom: 8px;
            border-bottom: 2px solid #426b9a;
            color: #426b9a;
            font-size: 19px;
        ">
            OTHER
        </h2>
        """

        for item in grouped["OTHER"]:

            summary = html_escape(
                item.get("Summary", "")
            ).strip()

            link = html_escape(
                item.get("Gmail Link", "")
            ).strip()

            if not summary and not link:
                continue

            html += f"""
            <div style="
                border: 1px solid #dddddd;
                border-radius: 8px;
                padding: 14px 16px;
                margin-bottom: 10px;
                background-color: #fafafa;
            ">

                <div style="
                    font-size: 14px;
                    line-height: 1.55;
                ">
                    {summary}
                </div>

                <div style="
                    margin-top: 9px;
                ">
                    <a
                        href="{link}"
                        style="
                            color: #426b9a;
                            font-size: 12px;
                            font-weight: bold;
                            text-decoration: none;
                        "
                    >
                        OPEN MAIL →
                    </a>
                </div>

            </div>
            """


    # =====================================================
    # FOOTER
    # =====================================================

    html += """
        <div style="
            margin-top: 30px;
            padding-top: 15px;
            border-top: 1px solid #eeeeee;
        ">

            <p style="
                color: #999999;
                font-size: 11px;
                margin: 0;
            ">
                College Mail Assistant
            </p>

        </div>

    </div>

    </body>

    </html>
    """

    return html


# =========================================================
# SEND EMAIL
# =========================================================

def send_summary_email(service, html):

    message = EmailMessage()

    message["To"] = MY_EMAIL

    message["Subject"] = (
        "College Mail Summary"
    )

    # Plain-text fallback
    message.set_content(
        """
College Mail Summary

Please open this email in an HTML-compatible
email client to view the formatted summary.
"""
    )

    # HTML version
    message.add_alternative(
        html,
        subtype="html"
    )

    # Encode MIME message
    encoded_message = base64.urlsafe_b64encode(
        message.as_bytes()
    ).decode()

    body = {
        "raw": encoded_message
    }

    result = (
        service.users()
        .messages()
        .send(
            userId="me",
            body=body
        )
        .execute()
    )

    print(
        "Summary email sent successfully!"
    )

    print(
        "Message ID:",
        result["id"]
    )


# =========================================================
# MAIN
# =========================================================

def main():

    print("Reading Excel...")

    records = read_excel()

    print(
        f"Total records: {len(records)}"
    )

    grouped = group_records(
        records
    )

    print(
        f"NEW COMPANY: "
        f"{len(grouped['NEW COMPANY'])}"
    )

    print(
        f"TEST + SHORTLIST: "
        f"{len(grouped['TEST + SHORTLIST'])}"
    )

    print(
        f"SELECTION: "
        f"{len(grouped['SELECTION'])}"
    )

    print(
        f"OTHER: "
        f"{len(grouped['OTHER'])}"
    )

    print("Creating HTML summary...")

    html = create_html(
        grouped
    )

    print("Connecting to Gmail...")

    service = get_gmail_service()

    print("Sending summary...")

    send_summary_email(
        service,
        html
    )


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":
    main()