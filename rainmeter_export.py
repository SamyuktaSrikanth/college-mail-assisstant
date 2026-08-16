from pathlib import Path
from openpyxl import load_workbook


# =========================================================
# PATHS
# =========================================================

PROJECT_DIR = Path(__file__).resolve().parent

EXCEL_FILE = (
    PROJECT_DIR
    / "data"
    / "College_Placement_2026.xlsx"
)

RAINMETER_FILE = Path(
    r"C:\Users\Samyukta\OneDrive\Documents"
    r"\Rainmeter\Skins\MailDashboard"
    r"\@Resources\mail_data.inc"
)



# =========================================================
# HELPERS
# =========================================================

def clean(value):
    if value is None:
        return ""

    return str(value).strip()


def escape(value):
    """
    Make Excel text safe for a Rainmeter variable.
    """

    value = clean(value)

    value = value.replace("\n", " ")
    value = value.replace("\r", " ")

    return value


# =========================================================
# READ EXCEL
# =========================================================

def read_excel():

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    records = []

    # Read every monthly sheet.
    for sheet in workbook.worksheets:

        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        headers = [
            clean(value)
            for value in rows[0]
        ]

        for row in rows[1:]:

            record = dict(
                zip(headers, row)
            )

            # Ignore completely empty rows.
            if not any(
                clean(value)
                for value in row
            ):
                continue

            records.append(record)

    return records


# =========================================================
# CLASSIFY FOR WIDGET
# =========================================================

def classify_records(records):

    data = {
        "new_company": [],
        "test_shortlist": [],
        "selection": [],
        "academic": []
    }

    for record in records:

        placement_type = clean(
            record.get("Placement Type")
        ).upper()

        summary = escape(
            record.get("Summary")
        )

        gmail_link = escape(
            record.get("Gmail Link")
        )

        # Ignore records without useful content.
        if not summary and not gmail_link:
            continue

        item = {
            "summary": summary,
            "gmail": gmail_link
        }

        # ---------------------------------------------
        # NEW COMPANY
        # ---------------------------------------------

        if placement_type == "NEW_COMPANY":

            data["new_company"].append(item)

        # ---------------------------------------------
        # TEST + SHORTLIST
        # ---------------------------------------------

        elif placement_type in {
            "SHORTLIST_TEST",
            "SHORTLISTED_TEST",
            "UPDATE"
        }:

            data["test_shortlist"].append(item)

        # ---------------------------------------------
        # SELECTION
        # ---------------------------------------------

        elif placement_type == "SELECTION":

            data["selection"].append(item)

        # ---------------------------------------------
        # ACADEMIC
        # OTHER → ACADEMIC FOR NOW
        # ---------------------------------------------

        elif placement_type == "OTHER":

            data["academic"].append(item)

    return data


# =========================================================
# WRITE RAINMETER FILE
# =========================================================

def write_rainmeter_file(data):

    lines = []

    lines.append("[Variables]")
    lines.append("")

    # -----------------------------------------------------
    # COUNTS
    # -----------------------------------------------------

    lines.append(
        f"NewCompanyCount="
        f"{len(data['new_company'])}"
    )

    lines.append(
        f"TestShortlistCount="
        f"{len(data['test_shortlist'])}"
    )

    lines.append(
        f"SelectionCount="
        f"{len(data['selection'])}"
    )

    lines.append(
        f"AcademicCount="
        f"{len(data['academic'])}"
    )

    lines.append("")


    # -----------------------------------------------------
    # NEW COMPANY
    # -----------------------------------------------------

    for i, item in enumerate(
        data["new_company"],
        start=1
    ):

        lines.append(
            f"NewSummary{i}="
            f"{item['summary']}"
        )

        lines.append(
            f"NewMail{i}="
            f"{item['gmail']}"
        )

        lines.append("")


    # -----------------------------------------------------
    # TEST + SHORTLIST
    # -----------------------------------------------------

    for i, item in enumerate(
        data["test_shortlist"],
        start=1
    ):

        lines.append(
            f"TestSummary{i}="
            f"{item['summary']}"
        )

        lines.append(
            f"TestMail{i}="
            f"{item['gmail']}"
        )

        lines.append("")


    # -----------------------------------------------------
    # SELECTION
    # -----------------------------------------------------

    for i, item in enumerate(
        data["selection"],
        start=1
    ):

        lines.append(
            f"SelectionSummary{i}="
            f"{item['summary']}"
        )

        lines.append(
            f"SelectionMail{i}="
            f"{item['gmail']}"
        )

        lines.append("")


    # -----------------------------------------------------
    # ACADEMIC
    # -----------------------------------------------------

    for i, item in enumerate(
        data["academic"],
        start=1
    ):

        lines.append(
            f"AcademicSummary{i}="
            f"{item['summary']}"
        )

        lines.append(
            f"AcademicMail{i}="
            f"{item['gmail']}"
        )

        lines.append("")


    # -----------------------------------------------------
    # WRITE FILE
    # -----------------------------------------------------

    RAINMETER_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    RAINMETER_FILE.write_text(
        "\n".join(lines),
        encoding="utf-8"
    )

    print(
        f"\nRainmeter data written to:\n"
        f"{RAINMETER_FILE}"
    )


# =========================================================
# MAIN
# =========================================================

def main():

    records = read_excel()

    print(
        f"Excel records found: {len(records)}"
    )

    data = classify_records(records)

    print(
        f"New Company: {len(data['new_company'])}"
    )

    print(
        f"Test + Shortlist: "
        f"{len(data['test_shortlist'])}"
    )

    print(
        f"Selection: {len(data['selection'])}"
    )

    print(
        f"Academic: {len(data['academic'])}"
    )

    write_rainmeter_file(data)


if __name__ == "__main__":
    main()